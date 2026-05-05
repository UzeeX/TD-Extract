# app.py
# TD Wealth Locator Extractor - Streamlit Cloud version
#
# Required repo files:
# app.py
# requirements.txt
#
# IMPORTANT:
# Do NOT use packages.txt.
# Do NOT use playwright install-deps.
# This app only uses: python -m playwright install chromium

import os
import re
import sys
import subprocess
from io import BytesIO
from urllib.parse import urljoin

import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# ----------------------------- Page Setup -----------------------------

st.set_page_config(page_title="TD Wealth Locator Extractor", layout="wide")

st.markdown(
    """
<style>
.block-container { max-width: 1180px; padding-top: 2rem; padding-bottom: 3rem; }
h1, h2, h3 { letter-spacing: -0.02em; }
.card {
  border: 1px solid rgba(0,0,0,0.10);
  background: rgba(255,255,255,0.70);
  border-radius: 18px;
  padding: 16px 18px;
  box-shadow: 0 6px 18px rgba(0,0,0,0.06);
}
.small-muted { color: rgba(0,0,0,0.55); font-size: 0.92rem; }
</style>
""",
    unsafe_allow_html=True,
)

st.title("TD Wealth Locator Extractor")
st.caption("Searches TD Wealth locator by category, province, and multiple cities. Exports CSV and optional Excel.")


# ----------------------------- Constants -----------------------------

TD_LOCATOR_URL = "https://advisors.td.com/locator/"

TD_CATEGORIES = {
    "Investment Advisor or Branch": "Find an Investment Advisor or Branch",
    "Portfolio Manager": "Find a Portfolio Manager",
    "Private Banker": "Find a Private Banker",
    "Financial Planner": "Find a Financial Planner",
    "Wealth Advisor": "Find a Wealth Advisor",
}

PROVINCE_OPTIONS = {
    "Alberta": "AB",
    "British Columbia": "BC",
    "Manitoba": "MB",
    "New Brunswick": "NB",
    "Newfoundland and Labrador": "NL",
    "Northwest Territories": "NT",
    "Nova Scotia": "NS",
    "Nunavut": "NU",
    "Ontario": "ON",
    "Prince Edward Island": "PE",
    "Quebec": "QC",
    "Saskatchewan": "SK",
    "Yukon": "YT",
}

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)

PHONE_RE = re.compile(
    r"\b(?:1[-\s]?)?\(?\d{3}\)?[-\s]?\d{3}[-\s]?\d{4}\b"
)

POSTAL_RE = re.compile(
    r"\b[ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z]\s?\d[ABCEGHJ-NPRSTV-Z]\d\b",
    re.I,
)


# ----------------------------- Playwright Setup -----------------------------

@st.cache_resource(show_spinner=False)
def install_playwright_chromium_once():
    """
    Installs Playwright Chromium once per Streamlit machine session.
    Do NOT use install-deps. That causes apt conflicts on Streamlit Cloud.
    """
    try:
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        return {
            "ok": True,
            "stdout": result.stdout,
            "stderr": result.stderr,
        }

    except subprocess.CalledProcessError as e:
        return {
            "ok": False,
            "stdout": e.stdout,
            "stderr": e.stderr,
        }


def launch_browser(playwright, slow_mo_ms: int):
    """
    Launch Playwright Chromium using the browser installed by:
    python -m playwright install chromium
    """
    try:
        return playwright.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--disable-setuid-sandbox",
                "--single-process",
                "--no-zygote",
                "--disable-extensions",
                "--disable-background-networking",
                "--disable-sync",
                "--disable-default-apps",
                "--disable-popup-blocking",
                "--disable-features=TranslateUI",
            ],
            slow_mo=slow_mo_ms,
        )

    except Exception as e:
        st.error("Chromium failed to launch on Streamlit Cloud.")
        st.write("The app installed Playwright Chromium, but the browser crashed when launching.")
        st.write("Do not add packages.txt. Do not use install-deps.")
        st.code(str(e))
        st.stop()


# ----------------------------- Helpers -----------------------------

def parse_city_filter(city_text: str) -> list[str]:
    return [c.strip() for c in city_text.split(",") if c.strip()]


def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D+", "", phone or "")

    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]

    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"

    return (phone or "").strip()


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def accept_cookies_if_present(page) -> None:
    cookie_texts = [
        "Accept all",
        "Accept All",
        "I Accept",
        "Accept",
        "Agree",
        "Continue",
        "Allow all",
    ]

    for txt in cookie_texts:
        try:
            page.get_by_text(txt, exact=False).first.click(timeout=1500)
            page.wait_for_timeout(500)
            return
        except Exception:
            continue


def click_text_if_exists(page, text: str, timeout: int = 4000) -> bool:
    try:
        page.get_by_text(text, exact=False).first.click(timeout=timeout)
        return True
    except Exception:
        return False


def fill_first_working_input(page, value: str) -> bool:
    selectors = [
        "input[type='search']",
        "input[placeholder*='Search']",
        "input[placeholder*='search']",
        "input[placeholder*='Location']",
        "input[placeholder*='location']",
        "input[aria-label*='Search']",
        "input[aria-label*='search']",
        "input[aria-label*='Location']",
        "input[aria-label*='location']",
        "input[type='text']",
    ]

    for selector in selectors:
        try:
            loc = page.locator(selector).first
            loc.wait_for(state="visible", timeout=5000)
            loc.fill("")
            loc.fill(value)
            page.wait_for_timeout(800)
            return True
        except Exception:
            continue

    return False


def press_search(page) -> None:
    try:
        page.keyboard.press("Enter")
        page.wait_for_timeout(2500)
    except Exception:
        pass

    possible_buttons = [
        "button:has-text('Search')",
        "input[type='submit']",
        "button[type='submit']",
        "[role='button']:has-text('Search')",
        "button:has-text('Find')",
        "[role='button']:has-text('Find')",
    ]

    for selector in possible_buttons:
        try:
            page.locator(selector).first.click(timeout=3000)
            page.wait_for_timeout(3000)
            return
        except Exception:
            continue


def auto_scroll(page, steps: int = 10, delay_ms: int = 600) -> None:
    for _ in range(steps):
        try:
            page.mouse.wheel(0, 1200)
            page.wait_for_timeout(delay_ms)
        except Exception:
            break


def extract_profile_links_from_html(html: str, base_url: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    links = []

    for a in soup.find_all("a", href=True):
        href = a.get("href", "").strip()
        text = clean_text(a.get_text(" ", strip=True))
        full_url = urljoin(base_url, href)

        if "advisors.td.com" not in full_url.lower():
            continue

        lower_url = full_url.lower()

        skip_patterns = [
            "/locator",
            "privacy",
            "legal",
            "accessibility",
            "login",
            "easyweb",
            "webbroker",
            "digitalvault",
            "contact-us",
            "terms",
        ]

        if any(skip in lower_url for skip in skip_patterns):
            continue

        links.append({
            "link_text": text,
            "profile_url": full_url,
        })

    seen = set()
    out = []

    for item in links:
        key = item["profile_url"]

        if key not in seen:
            seen.add(key)
            out.append(item)

    return out


def extract_result_cards_from_html(
    html: str,
    base_url: str,
    city: str,
    province_code: str,
    category: str,
) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    candidates = []

    possible_nodes = soup.find_all(["li", "article", "section", "div"])

    for node in possible_nodes:
        text = clean_text(node.get_text(" ", strip=True))

        if not text or len(text) < 20:
            continue

        if len(text) > 2500:
            continue

        has_phone = bool(PHONE_RE.search(text))
        has_postal = bool(POSTAL_RE.search(text))
        has_city = city.lower() in text.lower()

        a = node.find("a", href=True)
        profile_url = ""

        if a:
            profile_url = urljoin(base_url, a.get("href", "").strip())

        has_td_link = "advisors.td.com" in profile_url.lower()

        if not (has_phone or has_postal or has_city or has_td_link):
            continue

        emails = sorted(set(EMAIL_RE.findall(text)))

        phones = [normalize_phone(p) for p in PHONE_RE.findall(text)]
        phones = list(dict.fromkeys([p for p in phones if p]))

        lines = [
            clean_text(x)
            for x in node.get_text("\n", strip=True).split("\n")
            if clean_text(x)
        ]

        name = ""
        title = ""

        if lines:
            name = lines[0]

        if len(lines) > 1:
            title = lines[1]

        candidates.append({
            "category": category,
            "searched_city": city,
            "province": province_code,
            "name_or_branch": name,
            "title": title,
            "phone": " | ".join(phones[:3]),
            "email": emails[0] if emails else "",
            "address_hint": text[:500],
            "profile_url": profile_url,
        })

    if not candidates:
        return []

    df = pd.DataFrame(candidates)

    df["dedupe_key"] = (
        df["profile_url"].fillna("") + "|" +
        df["name_or_branch"].fillna("") + "|" +
        df["phone"].fillna("") + "|" +
        df["address_hint"].fillna("").str[:120]
    ).str.lower()

    df = df.drop_duplicates(subset=["dedupe_key"], keep="first")
    df = df.drop(columns=["dedupe_key"], errors="ignore")

    return df.to_dict("records")


def search_td_locator(
    category_label: str,
    category_button_text: str,
    city: str,
    province_label: str,
    province_code: str,
    slow_mo_ms: int,
    wait_seconds: float,
) -> tuple[list[dict], list[dict]]:
    rows = []
    errors = []

    install_status = install_playwright_chromium_once()

    if not install_status["ok"]:
        st.error("Playwright Chromium install failed.")
        st.write("Do not use packages.txt. Do not use install-deps.")
        st.code(install_status["stderr"] or install_status["stdout"])
        st.stop()

    search_location = f"{city}, {province_label}"

    with sync_playwright() as p:
        browser = launch_browser(p, slow_mo_ms=slow_mo_ms)

        context = browser.new_context(
            viewport={"width": 1440, "height": 1000},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-CA",
        )

        page = context.new_page()

        try:
            page.goto(TD_LOCATOR_URL, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(int(wait_seconds * 1000))

            accept_cookies_if_present(page)

            clicked_category = click_text_if_exists(page, category_button_text, timeout=7000)

            if not clicked_category:
                errors.append({
                    "city": city,
                    "step": "category",
                    "error": f"Could not click category: {category_button_text}",
                })

            page.wait_for_timeout(1500)

            click_text_if_exists(page, "By Location", timeout=3000)
            page.wait_for_timeout(800)

            filled = fill_first_working_input(page, search_location)

            if not filled:
                errors.append({
                    "city": city,
                    "step": "input",
                    "error": "Could not find/fill search input.",
                })

                html = page.content()

                rows = extract_result_cards_from_html(
                    html=html,
                    base_url=TD_LOCATOR_URL,
                    city=city,
                    province_code=province_code,
                    category=category_label,
                )

                browser.close()
                return rows, errors

            press_search(page)
            page.wait_for_timeout(int(wait_seconds * 1000))

            auto_scroll(page, steps=10, delay_ms=600)

            html = page.content()

            rows = extract_result_cards_from_html(
                html=html,
                base_url=TD_LOCATOR_URL,
                city=city,
                province_code=province_code,
                category=category_label,
            )

            if not rows:
                links = extract_profile_links_from_html(html, TD_LOCATOR_URL)

                for link in links:
                    rows.append({
                        "category": category_label,
                        "searched_city": city,
                        "province": province_code,
                        "name_or_branch": link.get("link_text", ""),
                        "title": "",
                        "phone": "",
                        "email": "",
                        "address_hint": "",
                        "profile_url": link.get("profile_url", ""),
                    })

        except Exception as e:
            errors.append({
                "city": city,
                "step": "search",
                "error": str(e),
            })

        finally:
            try:
                browser.close()
            except Exception:
                pass

    return rows, errors


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TD Results"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    align = Alignment(vertical="center", wrap_text=True)

    ws.append(list(df.columns))

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for row_i in range(2, ws.max_row + 1):
        for col_i in range(1, ws.max_column + 1):
            ws.cell(row=row_i, column=col_i).alignment = align

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))

        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 65)

    bio = BytesIO()
    wb.save(bio)

    return bio.getvalue()


# ----------------------------- UI Controls -----------------------------

st.markdown('<div class="card">', unsafe_allow_html=True)

c1, c2, c3, c4 = st.columns([1.25, 1.25, 1.25, 1.25])

with c1:
    category_label = st.selectbox(
        "TD category",
        options=list(TD_CATEGORIES.keys()),
        index=0,
    )

    category_button_text = TD_CATEGORIES[category_label]

with c2:
    province_label = st.selectbox(
        "Province / Territory",
        options=list(PROVINCE_OPTIONS.keys()),
        index=list(PROVINCE_OPTIONS.keys()).index("British Columbia"),
    )

    province_code = PROVINCE_OPTIONS[province_label]

with c3:
    city_text = st.text_input(
        "Cities - separate multiple cities with commas",
        placeholder="e.g., Richmond, Vancouver, Surrey, Burnaby",
        help="Enter multiple cities separated by commas.",
    )

    selected_cities = parse_city_filter(city_text)

    if selected_cities:
        st.caption(f"Searching: {', '.join(selected_cities)}")

with c4:
    wait_seconds = st.slider("Wait after search seconds", 2.0, 15.0, 6.0, 0.5)

    max_cities = st.number_input(
        "Max cities 0 = no limit",
        min_value=0,
        max_value=100,
        value=0,
        step=1,
    )

st.markdown("</div>", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

st.markdown('<div class="card">', unsafe_allow_html=True)

c5, c6, c7 = st.columns([1, 1, 1])

with c5:
    slow_mo_ms = st.slider(
        "Browser slow motion ms",
        0,
        1000,
        100,
        50,
        help="Keep this low on Streamlit Cloud.",
    )

with c6:
    do_excel = st.toggle(
        "Also generate Excel .xlsx",
        value=False,
        disabled=not OPENPYXL_OK,
    )

with c7:
    show_debug = st.toggle(
        "Show debug errors",
        value=True,
    )

st.markdown("</div>", unsafe_allow_html=True)

run = st.button("Run TD Search", use_container_width=True)

st.divider()

if not run:
    st.info("Choose category/province, enter one or more cities, then click **Run TD Search**.")
    st.stop()

if not selected_cities:
    st.warning("Enter at least one city. Example: Richmond, Vancouver, Surrey.")
    st.stop()

if max_cities and int(max_cities) > 0:
    selected_cities = selected_cities[: int(max_cities)]


# ----------------------------- Main Run -----------------------------

status = st.empty()
progress = st.progress(0)

metrics = st.columns(4)
metrics[0].metric("Cities", len(selected_cities))
metrics[1].metric("Processed", 0)
metrics[2].metric("Rows", 0)
metrics[3].metric("Errors", 0)

all_rows = []
all_errors = []

for idx, city in enumerate(selected_cities, start=1):
    status.info(f"Searching TD locator for {category_label} in {city}, {province_label}...")

    rows, errors = search_td_locator(
        category_label=category_label,
        category_button_text=category_button_text,
        city=city,
        province_label=province_label,
        province_code=province_code,
        slow_mo_ms=slow_mo_ms,
        wait_seconds=wait_seconds,
    )

    all_rows.extend(rows)
    all_errors.extend(errors)

    progress.progress(int((idx / len(selected_cities)) * 100))

    metrics[1].metric("Processed", idx)
    metrics[2].metric("Rows", len(all_rows))
    metrics[3].metric("Errors", len(all_errors))

status.success("TD search complete.")

if all_errors and show_debug:
    st.subheader("Error samples")
    st.dataframe(pd.DataFrame(all_errors), use_container_width=True)

if not all_rows:
    st.warning(
        "No rows were extracted. Try one city first, increase wait time to 8-10 seconds, "
        "or TD may be blocking/headless browser automation."
    )
    st.stop()

df = pd.DataFrame(all_rows)

final_cols = [
    "category",
    "searched_city",
    "province",
    "name_or_branch",
    "title",
    "phone",
    "email",
    "address_hint",
    "profile_url",
]

for col in final_cols:
    if col not in df.columns:
        df[col] = ""

df = df[final_cols].copy()

df["dedupe_key"] = (
    df["category"].fillna("") + "|" +
    df["searched_city"].fillna("") + "|" +
    df["name_or_branch"].fillna("") + "|" +
    df["phone"].fillna("") + "|" +
    df["profile_url"].fillna("")
).str.lower()

df = df.drop_duplicates(subset=["dedupe_key"], keep="first")
df = df.drop(columns=["dedupe_key"], errors="ignore")

st.subheader("Preview")
st.dataframe(df, use_container_width=True, height=560)

csv_bytes = df.to_csv(index=False).encode("utf-8-sig")

st.download_button(
    "Download CSV",
    data=csv_bytes,
    file_name="td_wealth_locator_results.csv",
    mime="text/csv",
    use_container_width=True,
)

if do_excel and OPENPYXL_OK:
    try:
        xlsx_bytes = df_to_excel_bytes(df)

        st.download_button(
            "Download Excel .xlsx",
            data=xlsx_bytes,
            file_name="td_wealth_locator_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as e:
        st.warning(f"Excel export failed: {e}")

with st.expander("Setup notes"):
    st.write(
        "- Your GitHub repo should have only `app.py` and `requirements.txt`.\n"
        "- Delete `packages.txt` completely.\n"
        "- Do not use `playwright install-deps chromium`.\n"
        "- This app only runs `python -m playwright install chromium`.\n"
        "- Enter multiple cities separated by commas, like Richmond, Vancouver, Surrey.\n"
        "- First run may take 1-2 minutes while Chromium installs.\n"
        "- If Chromium still fails to launch, Streamlit Cloud may not support this Playwright setup reliably."
    )
